# フォルダ内のExcelファイルの内容を1ファイルにまとめる。
import openpyxl
import glob, os
 
# 定数
IN_F_NAME= '/Users/takemura/Desktop/meta/*.xlsx'            # 取りまとめるファイルが配置されたフォルダ
OUT_NAME= 'output.xlsx'
 
# ワークブックを新規作成する。
book = openpyxl.Workbook()
 
# シートの取得
sheet = book.active
 
work_row = 1                              # 現在作業中の行
# フォルダ内のファイルを取得
for gb in glob.glob(IN_F_NAME, recursive=True):
    # 取得したファイル名のプロンプト表示
    # print(gb)
 
    # ワークシート取得 （worksheets[0]で最初のシートを指定）
    sheet2 = openpyxl.load_workbook(gb).worksheets[0]
    
    # 取得したワークシートの最大行、最大列を取得
    max_column = sheet2.max_column
    max_row = sheet2.max_row
 
    # 取得したシートの最大行をプロンプト表示
    # print('max_row:' + str(max_row))
    
    # セル内容取得（'r in range(1'、'c in range(1' の部分でそれぞれシートの1行目、1列目からの取得を示す。）
    for r in range(1, max_row+1):
        for c in range(1, max_column+1):
             # セル内容のコピー
             sheet.cell(row=work_row,column=c).value = sheet2.cell(row=r,column=c).value
        work_row+=1
        
# ファイル保存（プログラム実行したフォルダに新規にファイルが作成）
book.save(os.path.dirname(os.path.abspath(__file__)) + '_' + OUT_NAME)
 
